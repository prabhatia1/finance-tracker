"""Create test report Excel for Finance Tracker."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Report"

# Styling
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F4858", end_color="2F4858", fill_type="solid")
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_align = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:H1')
ws['A1'] = "Finance Tracker — Full Feature Test Report"
ws['A1'].font = Font(bold=True, size=16, color="2F4858")
ws['A1'].alignment = Alignment(horizontal="center")

ws.merge_cells('A2:H2')
ws['A2'] = "Date: 2026-07-08 | Tester: Pratik Bhatia | App: Flask + SQLite + OpenPyXL"
ws['A2'].font = Font(size=10, color="666666")
ws['A2'].alignment = Alignment(horizontal="center")

# Headers
headers = ["#", "Feature", "Test Case", "Steps", "Expected Result", "Actual Result", "Status", "Severity / Notes"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Column widths
widths = [5, 18, 30, 48, 28, 32, 10, 32]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+i)].width = w

# Test data
tests = [
    ("Dashboard", "Page loads correctly",
     "1. Navigate to http://127.0.0.1:5000/",
     "Dashboard loads with Today's Transactions, Recent Transactions, toolbar buttons",
     "Loaded. Shows 6 today's txns, 20 recent txns, all toolbar buttons present",
     "PASS", ""),
    ("Dashboard", "Today's Transactions display",
     "1. Check Today's section shows only current date transactions",
     "Shows only 2026-07-08 transactions with description, amount, card, edit/delete",
     "Shows 6 transactions dated 2026-07-08 with all columns. Contains test data entries (ln, uyyu, ad)",
     "PASS", "Test data present but display works"),
    ("Dashboard", "Recent Transactions display",
     "1. Scroll to Recent Transactions section",
     "Shows 20 most recent txns across all dates, sorted by date desc",
     "Shows 20 transactions sorted by date descending with edit links",
     "PASS", ""),
    ("Dashboard", "Edit button navigates to edit page",
     "1. Click pencil icon on any transaction",
     "Navigates to /edit/<id> with pre-filled form",
     "Navigated to edit page with all fields pre-filled correctly (date, type, amount, desc, category, card, person)",
     "PASS", ""),
    ("Dashboard", "Delete button shows confirmation",
     "1. Click trash icon on any transaction",
     "Browser shows confirm dialog before deleting",
     "Confirm dialog appears, transaction removed after OK",
     "PASS", "JS confirm dialog"),
    ("Dashboard", "Toolbar buttons functional",
     "1. Check Add Transaction, Upload, Reports, Sync Excel, Open Excel, Export CSV buttons",
     "All buttons navigate to correct pages or trigger actions",
     "All 6 toolbar buttons present and clickable",
     "PASS", ""),
    ("Navigation", "All nav links work and highlight active",
     "1. Click each nav link and verify page loads",
     "Each link navigates to correct page, active link highlighted",
     "All 7 nav items (Dashboard, Add, Upload, Reports, Settings, People, Balances) work",
     "PASS", ""),

    ("Add Transaction", "Form renders with all required fields",
     "1. Navigate to /add",
     "Shows: Date, Type, Amount, Description, Category, Card*(required), Notes, Person fields",
     "All fields present. Card has required-validation, Person dropdown has 7 people",
     "PASS", ""),
    ("Add Transaction", "Card selection is required",
     "1. Leave card as '-- Select Card --' and submit",
     "Form validation prevents submission until card is selected",
     "Card field has 'required' attribute; submission blocked without selection",
     "PASS", ""),
    ("Add Transaction", "Person dropdown correct",
     "1. Open person dropdown",
     "Shows: -- None --, Mom, Dad, Brother, Sister, Ravi, Amit, Priya",
     "All 7 people listed with -- None -- option",
     "PASS", ""),
    ("Add Transaction", "Auto-categorization works",
     "1. Enter 'swiggy zomato' in description\n2. Select Auto-detect category\n3. Submit",
     "Description keywords auto-assign to matching category (Dining / Food)",
     "Keywords like 'swiggy', 'zomato' correctly map to Dining/Food",
     "PASS", ""),
    ("Add Transaction", "Submit creates transaction in DB",
     "1. Fill all fields, submit\n2. Check dashboard",
     "Transaction appears in Today's Transactions and Recent list",
     "Transaction added successfully and visible on dashboard",
     "PASS", ""),

    ("Edit Transaction", "Form pre-fills existing data",
     "1. Click edit on a transaction",
     "All fields pre-filled: date, type, amount, desc, category, card, notes, person",
     "All 8 fields pre-filled correctly from DB",
     "PASS", ""),
    ("Edit Transaction", "Update saves changes",
     "1. Change description, click Update",
     "Flash message confirms, dashboard shows updated description",
     "Changes saved and reflected on dashboard immediately",
     "PASS", ""),

    ("Delete Transaction", "Confirmation before delete",
     "1. Click delete on transaction",
     "Browser confirm dialog appears",
     "Confirm dialog shown",
     "PASS", ""),
    ("Delete Transaction", "Transaction removed from DB",
     "1. Confirm deletion",
     "Transaction removed from DB and dashboard, flash message shown",
     "Removed successfully, no longer in transactions list",
     "PASS", ""),

    ("Reports", "Page loads with all sections",
     "1. Navigate to /reports?month=7&year=2026&card=all",
     "Shows Daily Summary, Category Breakdown, Card-wise Spending",
     "All 3 sections display. Daily Summary (6 txns today, ₹12,645), Category Breakdown (20 txns, ₹50,783), Card-wise",
     "PASS", ""),
    ("Reports", "Month filter works",
     "1. Change month dropdown to another month",
     "Reports update for selected month",
     "Month filter works, no data months show empty",
     "PASS", ""),
    ("Reports", "Card filter works",
     "1. Select specific card from dropdown",
     "Reports show only that card's transactions",
     "Filter by SBI Cashback shows 2 txns (₹715), HDFC Millennia shows 1 txn (₹15,000)",
     "PASS", ""),
    ("Reports", "Export Month CSV button present",
     "1. Check for Export CSV link",
     "Link downloads CSV for current month",
     "Export Month button present",
     "PASS", "Not fully verified (no download in test)"),

    ("People", "Page shows correct balances",
     "1. Navigate to /people",
     "Shows all people with totals: you paid, they paid back, net balance",
     "Priya: owes ₹11,930 (4 txns), Ravi: owes ₹515 (1 txn)",
     "PASS", ""),
    ("People", "Filter by person works",
     "1. Select Priya from filter dropdown",
     "Only Priya's transactions shown, balance card updates",
     "Filter shows 4 Priya transactions correctly",
     "PASS", ""),
    ("People", "Clear filter works",
     "1. Apply filter, then click 'Clear' link",
     "Returns to showing all people",
     "Clear resets to -- All People -- view",
     "PASS", ""),
    ("People", "DEFECT: '1 txns' grammar",
     "1. Check Ravi row (1 transaction) in people table",
     "Should show '1 txn' (singular), not '1 txns'",
     "Shows '1 txns' which is grammatically wrong",
     "FAIL", "MINOR: Use '1 txn' for singular"),

    ("Balances", "All 12 months displayed",
     "1. Navigate to /balances",
     "Shows 12 rows (Jan-Dec) with start/end balance inputs + calculated fields",
     "12 months shown, July has data (start=50000, debit=50783, expected=-783), rest zero",
     "PASS", ""),
    ("Balances", "Start balance saves to DB on blur",
     "1. Enter 50000 in July Start Balance\n2. Tab out",
     "AJAX save to DB, Expected End auto-recalculates",
     "Start=₹50,000 saved. Expected End = ₹-783 (50000-50783)",
     "PASS", ""),
    ("Balances", "End balance saves on blur",
     "1. Enter end balance field\n2. Tab out",
     "Value saved, Difference column calculates (actual - expected)",
     "End balance saves correctly via POST",
     "PASS", ""),
    ("Balances", "Net Change and Difference auto-calc",
     "1. View July: net=-50783, expected=-783, actual=0",
     "Net=Debit-Credit, Expected=Start+Net, Diff=Actual-Expected",
     "All calculated correctly server-side",
     "PASS", ""),

    ("Settings", "Cards CRUD",
     "1. View cards table, add card, remove card",
     "All 6 cards shown, add creates new card, remove deletes with confirmation",
     "6 cards present (SBI Cashback, SBI PhonePe, HDFC Millennia, HDFC Swiggy, BOB Eterna, Other Card). Add/remove work",
     "PASS", ""),
    ("Settings", "Categories CRUD",
     "1. View categories table with keyword rules",
     "14 categories shown with keywords, type, delete option",
     "Categories display correctly with keyword chips and type badges",
     "PASS", ""),
    ("Settings", "People balance column displayed",
     "1. Check People section table",
     "Shows Name, Balance (with colored badge), Delete button",
     "Priya: 'Owes you ₹11,930'(green), Ravi: 'Owes you ₹515'(green), others: '--'",
     "PASS", "New feature verified"),
    ("Settings", "Delete person blocked (has balance)",
     "1. Try to delete Priya who owes ₹11,930",
     "Flash error: 'Cannot remove Priya -- pending amount... Settle up first!'",
     "Deletion blocked with appropriate flash error message",
     "PASS", "New feature verified"),
    ("Settings", "Delete person allowed (zero balance)",
     "1. Try to delete Mom (no transactions)",
     "Person removed without error",
     "Deleted successfully (zero balance)",
     "PASS", ""),

    ("Excel Sync", "DB to Excel sync runs without error",
     "1. Run db_sync_to_excel()",
     "All DB transactions written to correct category cells in month sheets",
     "17 of 20 transactions synced. Missing: ln x2, uyyu, ad (test data with 'Other' category)",
     "FAIL", "CRITICAL: Test data not written to correct Excel rows"),
    ("Excel Sync", "Card data preserved in Transactions sheet",
     "1. Check Card column in Transactions sheet after sync",
     "Card IDs should match DB (e.g. hdfc_mil for Rent, sbi_cb for SBI Cashback txns)",
     "ALL rows show card='other' regardless of actual card in DB",
     "FAIL", "CRITICAL: read_transactions_from_excel hardcodes card_id='other'"),
    ("Excel Sync", "Balances synced to Excel",
     "1. Check July sheet Row 2 (F2) and Row 38 (F38)",
     "Start Balance (F2) and End Balance (F38) written to Excel",
     "F2=50000 (Start Balance), F38=0 (End Balance) -- correct",
     "PASS", ""),
    ("Excel Sync", "Excel to DB sync (smart_sync)",
     "1. Add transaction in Excel, click Sync Excel button",
     "Excel data imported to DB without duplicates",
     "Smart sync detects changes and imports correctly",
     "PASS", ""),
    ("Excel Sync", "Category: 'Other' maps to wrong Excel row",
     "1. Check WEBCAT_TO_EXCEL priority for 'Other' category",
     "Should default to 'Misc Major' or 'Misc Daily' for unmatched descriptions",
     "'Other' web cat maps to 'Insurance (Other)' first -- causes test data to appear as insurance",
     "FAIL", "MEDIUM: Priority order bug in WEBCAT_TO_EXCEL['Other']"),

    ("Excel Integrity", "All 14 sheets present",
     "1. Check sheet names",
     "14 sheets: start, January-December, Transactions",
     "start, Jan-Dec, Transactions -- all 14 present",
     "PASS", ""),
    ("Excel Integrity", "Day row format correct",
     "1. Check Row 5 in July sheet",
     "Day labels as text ('01-Jul', '02-Jul'...) in D-AH",
     "Row 5: D5=01-Jul, E5=02-Jul, F5=03-Jul... correct text format",
     "PASS", ""),
    ("Excel Integrity", "Balance rows in all month sheets",
     "1. Check all months for Row 2 (F2) and Row 38 (F38)",
     "Each month has Start Balance and End Balance rows",
     "All 12 months have balance rows. Only July has data (50000 start, 0 end)",
     "PASS", ""),
    ("Excel Integrity", "Orphan data: Insurance (Other) ₹12,245",
     "1. Compare Excel data with DB transactions",
     "No data in Excel without corresponding DB entry",
     "Insurance (Other) row has ₹12,245 on day 8 -- accumulated from test data, not an actual insurance txn",
     "FAIL", "MEDIUM: Excel has orphan data from poor category mapping"),
    ("Excel Integrity", "Transactions sheet incomplete",
     "1. Compare Transactions sheet row count with DB",
     "All 20 DB entries should appear",
     "Only 18 rows in Transactions sheet. Missing 2 test entries",
     "FAIL", "MEDIUM: Incomplete sync due to card mapping issues"),

    ("Data Quality", "No test/garbage data in DB",
     "1. Query DB for meaningless descriptions",
     "All transactions should have real meaningful descriptions",
     "4 garbage entries: 'ln' x2 (₹11,230), 'uyyu' (₹515), 'ad' (₹500). Total ₹12,245 noise",
     "FAIL", "MEDIUM: Test data pollutes real records"),
    ("Data Quality", "No duplicate transactions",
     "1. Check for identical date+desc+amount+person combos",
     "No duplicate entries",
     "'ln' transaction appears TWICE (IDs #1223, #1224) -- same date, amount, person",
     "FAIL", "MEDIUM: Duplicate ₹5,615 entry for Priya"),
]

# Write test data
for i, (feature, test_case, steps, expected, actual, status, notes) in enumerate(tests, 1):
    row = i + 4
    ws.cell(row=row, column=1, value=i).alignment = body_align
    ws.cell(row=row, column=2, value=feature).alignment = body_align
    ws.cell(row=row, column=3, value=test_case).alignment = body_align
    ws.cell(row=row, column=4, value=steps).alignment = body_align
    ws.cell(row=row, column=5, value=expected).alignment = body_align
    ws.cell(row=row, column=6, value=actual).alignment = body_align
    ws.cell(row=row, column=7, value=status).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=row, column=8, value=notes).alignment = body_align

    # Color status cell
    status_cell = ws.cell(row=row, column=7)
    if status == "PASS":
        status_cell.fill = pass_fill
    elif status == "FAIL":
        status_cell.fill = fail_fill
    else:
        status_cell.fill = warn_fill

    # Borders
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).font = Font(size=9)

# Summary section
summary_row = len(tests) + 6
summary_row += 1
ws.merge_cells(f'A{summary_row}:H{summary_row}')
ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=14, color="2F4858")

pass_count = sum(1 for t in tests if t[5] == "PASS")
fail_count = sum(1 for t in tests if t[5] == "FAIL")
total = len(tests)

summary_row += 1
for label, val, color in [
    (f"Total Tests: {total}", None, None),
    (f"PASSED: {pass_count} ({pass_count*100//total}%)", pass_fill, "006100"),
    (f"FAILED: {fail_count} ({fail_count*100//total}%)", fail_fill, "9C0006"),
]:
    ws.cell(row=summary_row, column=1, value=label).font = Font(bold=True, size=11, color=color or "000000")
    if val:
        ws.cell(row=summary_row, column=1).fill = val
    summary_row += 1

# Defect section
summary_row += 1
ws.merge_cells(f'A{summary_row}:H{summary_row}')
ws.cell(row=summary_row, column=1, value="DEFECTS FOUND").font = Font(bold=True, size=14, color="9C0006")
ws.cell(row=summary_row, column=1).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

defects = [
    ("CRITICAL", "Card data lost in Excel round-trip",
     "read_transactions_from_excel() hardcodes card_id='other' at line 299. DB-to-Excel sync preserves card_id in month sheets but NOT in flat Transactions sheet. All cards revert to 'other'."),
    ("CRITICAL", "Test data not syncing to monthly sheet rows",
     "Transactions with web category 'Other' don't write to Misc Major/Misc Daily rows. 3 of 20 DB entries invisible in month sheets after full sync."),
    ("MEDIUM", "Category mapping: 'Other' -> 'Insurance (Other)' takes priority",
     "WEBCAT_TO_EXCEL['Other'] lists Insurance (Other) first. Unmatched 'Other' transactions land there instead of Misc Major. Defaut should be Misc Major for large amounts, Misc Daily for small."),
    ("MEDIUM", "Test/garbage data in production DB",
     "4 entries with meaningless descriptions: 'ln' x2 (₹11,230, person=Priya), 'uyyu' (₹515, person=Ravi, card=sbi_cb), 'ad' (₹500, person=Priya, card=sbi_pp). Total ₹12,245 noise."),
    ("MEDIUM", "Duplicate transaction in DB",
     "Transaction IDs #1223 and #1224 both: date=2026-07-08, description='ln', amount=₹5,615, category='Other', person='Priya'. Likely accidental double-submit."),
    ("MEDIUM", "Excel orphan data in Insurance (Other)",
     "Insurance (Other) row shows ₹12,245 accumulated from incorrectly mapped test data. Excel has data that doesn't represent an actual insurance transaction."),
    ("MINOR", "Pluralization: '1 txns'",
     "People page and card stats show '1 txns' instead of '1 txn' for singular transaction count."),
]

for i, (sev, title, detail) in enumerate(defects, 1):
    summary_row += 1
    ws.cell(row=summary_row, column=1, value=i).alignment = body_align
    ws.cell(row=summary_row, column=2, value=sev).alignment = body_align
    ws.cell(row=summary_row, column=3, value=title)
    ws.merge_cells(f'D{summary_row}:H{summary_row}')
    ws.cell(row=summary_row, column=4, value=detail).alignment = body_align

    sev_cell = ws.cell(row=summary_row, column=2)
    if sev == "CRITICAL":
        sev_cell.fill = fail_fill
    elif sev == "MEDIUM":
        sev_cell.fill = warn_fill

    for col in range(1, 9):
        ws.cell(row=summary_row, column=col).border = thin_border
        ws.cell(row=summary_row, column=col).font = Font(size=9)

# Freeze panes + auto-filter
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:H{len(tests)+4}"

# Row heights
for r in range(5, len(tests)+5):
    ws.row_dimensions[r].height = 55
ws.row_dimensions[4].height = 30

wb.save('test_report.xlsx')
print(f"✅ test_report.xlsx created!")
print(f"   {total} test cases ({pass_count} PASS / {fail_count} FAIL)")
print(f"   {len(defects)} defects documented")
