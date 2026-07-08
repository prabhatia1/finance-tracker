#!/usr/bin/env python3
"""Create a fresh Excel with monthly sheets (calendar layout) like the original."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

YEAR = 2026

# ─── Categories from your actual setup ───
MAJOR_EXPENSES = [
    "Rent / Home Loan",
    "Electricity Bill",
    "LIC Premium",
    "EMI / Loan Payment",
    "School / Education",
    "Insurance (Other)",
    "Property Tax / Maintenance",
    "Internet / DTH Recharge",
    "Groceries (Bulk)",
    "Medical / Health",
    "Travel / Vacation",
    "Shopping (Big)",
    "Family Support",
    "Misc Major",
]

DAILY_EXPENSES = [
    "Groceries / Kirana",
    "Dining / Food / Swiggy",
    "Fuel / Petrol",
    "Recharge / Mobile",
    "Entertainment / OTT",
    "Transport / Uber / Auto",
    "Tea / Coffee / Snacks",
    "Household Items",
    "Personal Care",
    "Misc Daily",
]

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
CATEGORY_FILL = PatternFill("solid", fgColor="D9E2F3")
TODAY_FILL = PatternFill("solid", fgColor="FFE699")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
INCOME_FILL = PatternFill("solid", fgColor="C6EFCE")

# ─── Income Sources ───
INCOME_SOURCES = [
    "Salary / Main Income",
    "Freelance / Side Income",
    "Rental Income",
    "Investments / Dividends",
    "Other Income",
]


def create_monthly_excel(path):
    wb = openpyxl.Workbook()

    # ── Start Sheet ──
    ws_start = wb.active
    ws_start.title = "start"
    ws_start["D3"] = "Year :"
    ws_start["E3"] = YEAR
    ws_start["D3"].font = Font(bold=True, size=12)
    ws_start["E3"].font = Font(bold=True, size=14, color="4472C4")

    ws_start["D7"] = "Major Expenses:"
    ws_start["D7"].font = Font(bold=True, size=11, color="C00000")
    ws_start["F7"] = "Day-To-Day Expenses:"
    ws_start["F7"].font = Font(bold=True, size=11, color="C00000")
    ws_start["H7"] = "Income sources:"
    ws_start["H7"].font = Font(bold=True, size=11, color="006100")

    for i, name in enumerate(MAJOR_EXPENSES):
        row = 11 + i
        ws_start[f"D{row}"] = name
        ws_start[f"D{row}"].font = Font(size=10)
        ws_start[f"E{row}"] = 0  # Budget amount
        ws_start[f"E{row}"].number_format = '#,##0'

    for i, name in enumerate(DAILY_EXPENSES):
        row = 11 + i
        ws_start[f"F{row}"] = name
        ws_start[f"F{row}"].font = Font(size=10)
        ws_start[f"G{row}"] = 0
        ws_start[f"G{row}"].number_format = '#,##0'

    for i, name in enumerate(INCOME_SOURCES):
        row = 11 + i
        ws_start[f"H{row}"] = name
        ws_start[f"H{row}"].font = Font(size=10)
        ws_start[f"I{row}"] = 0
        ws_start[f"I{row}"].number_format = '#,##0'

    ws_start.column_dimensions["D"].width = 30
    ws_start.column_dimensions["E"].width = 15
    ws_start.column_dimensions["F"].width = 30
    ws_start.column_dimensions["G"].width = 15
    ws_start.column_dimensions["H"].width = 30
    ws_start.column_dimensions["I"].width = 15

    # ── Month Sheets ──
    day_headers = list(range(1, 32))

    import calendar
    for midx, month_name in enumerate(MONTHS):
        ws = wb.create_sheet(month_name)
        num_days = 31  # We'll use 31 columns, some stay empty for shorter months
        days_in_month = calendar.monthrange(YEAR, midx + 1)[1]

        # Row 1: Month selector / index
        ws.cell(row=1, column=1, value=f"=MATCH(B2,$BD$1:$BD$12,0)")

        # Row 2: Month name, Balance
        ws.cell(row=2, column=2, value=month_name)
        ws.cell(row=2, column=2).font = Font(bold=True, size=14, color="4472C4")
        ws.cell(row=2, column=4, value="=start!E3")
        ws.cell(row=2, column=5, value=f"Start Balance:")
        bal_col = get_column_letter(32 + 1)  # Column AF
        ws.cell(row=2, column=6, value=37561)  # Starting balance placeholder

        # Row 3: Empty
        # Row 4: Empty
        # Row 5: Day date as text (01-Jul, 02-Jul...)
        for day in range(1, num_days + 1):
            col = day + 3  # D=4, so day 1 = col 4 (D)
            cl = get_column_letter(col)

            # Row 5: Day date as text
            if day <= days_in_month:
                dt = date(YEAR, midx+1, day)
                cell_r5 = ws.cell(row=5, column=col, value=dt.strftime('%d-%b'))
            else:
                cell_r5 = ws.cell(row=5, column=col)
            cell_r5.number_format = '@'
            cell_r5.font = Font(bold=True, size=10)
            cell_r5.alignment = Alignment(horizontal="center")
            cell_r5.fill = HEADER_FILL
            cell_r5.font = HEADER_FONT

            # Wider columns for date text
            ws.column_dimensions[cl].width = 8.5

        ws.row_dimensions[5].height = 18

        # Row 5, col B: "Budgeted" label
        ws.cell(row=5, column=3, value="Budgeted").font = Font(bold=True, size=10)

        # ── Major Expenses rows (6 to 6+N) ──
        start_row = 6
        for i, name in enumerate(MAJOR_EXPENSES):
            row = start_row + i
            ref_row = 11 + i
            # Column A: formula reference to start sheet
            ws.cell(row=row, column=1, value=f"=start!D{ref_row}")
            # Column B: actual category name (readable in openpyxl)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=2).font = Font(size=9, bold=True)
            ws.cell(row=row, column=2).fill = CATEGORY_FILL
            # Column C: Budgeted amount
            ws.cell(row=row, column=3, value=f"=start!E{ref_row}")
            ws.cell(row=row, column=3).number_format = '#,##0'
            # Date columns
            for day in range(1, num_days + 1):
                col = day + 3
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0'
                cell.border = THIN_BORDER
            # Total (col 35 = AI)
            tot_col = get_column_letter(35)
            ws.cell(row=row, column=35,
                    value=f"=SUM(D{row}:AH{row})")
            ws.cell(row=row, column=35).number_format = '#,##0'
            ws.cell(row=row, column=35).fill = TOTAL_FILL
            ws.cell(row=row, column=35).font = Font(bold=True)
            # Count
            ws.cell(row=row, column=36,
                    value=f'=COUNTIF(D{row}:AH{row},">0")')
            ws.cell(row=row, column=36).font = Font(size=9)
            # Average
            ws.cell(row=row, column=37,
                    value=f'=IF(SUM(D{row}:AH{row})>0,AVERAGE(D{row}:AH{row}),)')
            ws.cell(row=row, column=37).number_format = '#,##0'
            # Actual
            ws.cell(row=row, column=38, value=f"=C{row}-AI{row}")
            ws.cell(row=row, column=38).number_format = '#,##0'
            # Next month link
            next_m = MONTHS[(midx + 1) % 12]
            ws.cell(row=row, column=58, value=next_m)

        # ── Day-To-Day Expenses rows ──
        dtd_start = start_row + len(MAJOR_EXPENSES)
        # Label row
        ws.cell(row=dtd_start, column=1, value="Day-To-Day Expenses")
        ws.cell(row=dtd_start, column=1).font = Font(bold=True, size=10, color="C00000")

        for i, name in enumerate(DAILY_EXPENSES):
            row = dtd_start + 1 + i
            ref_row = 11 + i
            ws.cell(row=row, column=1, value=f"=start!F{ref_row}")
            # Column B: actual category name
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=2).font = Font(size=9, bold=True)
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="FCE4D6")
            ws.cell(row=row, column=3, value=f"=start!G{ref_row}")
            ws.cell(row=row, column=3).number_format = '#,##0'
            for day in range(1, num_days + 1):
                col = day + 3
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0'
                cell.border = THIN_BORDER
            ws.cell(row=row, column=35,
                    value=f"=SUM(D{row}:AH{row})")
            ws.cell(row=row, column=35).number_format = '#,##0'
            ws.cell(row=row, column=35).fill = TOTAL_FILL
            ws.cell(row=row, column=35).font = Font(bold=True)
            ws.cell(row=row, column=36,
                    value=f'=COUNTIF(D{row}:AH{row},">0")')
            ws.cell(row=row, column=36).font = Font(size=9)
            ws.cell(row=row, column=37,
                    value=f'=IF(SUM(D{row}:AH{row})>0,AVERAGE(D{row}:AH{row}),)')
            ws.cell(row=row, column=37).number_format = '#,##0'
            ws.cell(row=row, column=38, value=f"=C{row}-AI{row}")
            ws.cell(row=row, column=38).number_format = '#,##0'
            next_m = MONTHS[(midx + 1) % 12]
            ws.cell(row=row, column=58, value=next_m)

        # ── Income rows ──
        income_start = dtd_start + 1 + len(DAILY_EXPENSES) + 1
        ws.cell(row=income_start - 1, column=1, value="Income")
        ws.cell(row=income_start - 1, column=1).font = Font(bold=True, size=10, color="006100")

        for i, name in enumerate(INCOME_SOURCES):
            row = income_start + i
            ref_row = 11 + i
            ws.cell(row=row, column=1, value=f"=start!H{ref_row}")
            # Column B: actual income source name
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=2).font = Font(size=9, bold=True)
            ws.cell(row=row, column=2).fill = INCOME_FILL
            ws.cell(row=row, column=3, value=f"=start!I{ref_row}")
            ws.cell(row=row, column=3).number_format = '#,##0'
            for day in range(1, num_days + 1):
                col = day + 3
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0'
                cell.border = THIN_BORDER
            ws.cell(row=row, column=35,
                    value=f"=SUM(D{row}:AH{row})")
            ws.cell(row=row, column=35).number_format = '#,##0'
            ws.cell(row=row, column=35).fill = TOTAL_FILL
            ws.cell(row=row, column=35).font = Font(bold=True)
            ws.cell(row=row, column=36,
                    value=f'=COUNTIF(D{row}:AH{row},">0")')
            ws.cell(row=row, column=36).font = Font(size=9)
            ws.cell(row=row, column=37,
                    value=f'=IF(SUM(D{row}:AH{row})>0,AVERAGE(D{row}:AH{row}),)')
            ws.cell(row=row, column=37).number_format = '#,##0'
            ws.cell(row=row, column=38, value=f"=C{row}-AI{row}")
            ws.cell(row=row, column=38).number_format = '#,##0'

        # ── Grand Total row ──
        grand_row = income_start + len(INCOME_SOURCES) + 1
        ws.cell(row=grand_row, column=1, value="Grand Total Expenses")
        ws.cell(row=grand_row, column=1).font = Font(bold=True, size=11)
        ws.cell(row=grand_row, column=35,
                value=f"=SUM(AI{start_row}:AI{start_row+len(MAJOR_EXPENSES)-1})+SUM(AI{dtd_start+1}:AI{dtd_start+len(DAILY_EXPENSES)})")
        ws.cell(row=grand_row, column=35).font = Font(bold=True, size=11)

        # Column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["AI"].width = 10
        ws.column_dimensions["AJ"].width = 7
        ws.column_dimensions["AK"].width = 9
        ws.column_dimensions["AL"].width = 10

        # Freeze panes (header rows + day columns)
        ws.freeze_panes = "D6"

    # ── Remove default sheet if extra ──
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(str(path))
    print(f"✅ Created: {path}")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Year: {YEAR}")
    print(f"   Major Expenses: {len(MAJOR_EXPENSES)} categories")
    print(f"   Daily Expenses: {len(DAILY_EXPENSES)} categories")
    print(f"   Income Sources: {len(INCOME_SOURCES)} sources")
    wb.close()


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\prati\finance-tracker\expense_tracker.xlsx"
    create_monthly_excel(out)
