"""
📊 Excel Sync Engine — Reads/Writes monthly calendar sheets.

Format:
  - 'start' sheet: config (year, budget categories)
  - 'January'..'December': calendar matrix (categories × days 1-31)
    - Row 5: day numbers (1-31) in columns D-AH
    - Rows 6+: categories, amounts entered in day columns
  - 'Transactions' sheet: auto-generated flat view for website
"""

import os
import re
import copy
from datetime import datetime, date
from pathlib import Path
from shutil import copy2

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    # Stub openpyxl types so the rest of the module can reference them
    Font = object
    PatternFill = object
    Alignment = object
    Border = object
    Side = object
    get_column_letter = lambda c: c

BASE_DIR = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "expense_tracker.xlsx"

# Same category order as create_monthly_excel.py
MAJOR_EXPENSES = [
    "Rent / Home Loan",
    "Electricity Bill",
    "LIC Premium",
    "EMI / Loan Payment",
    "School / Education",
    "Insurance (Other)",
    "Property Tax / Maintenance",
    "Internet / DTH Recharge",
    "Groceries (Bulk)",
    "Medical / Health",
    "Travel / Vacation",
    "Shopping (Big)",
    "Family Support",
    "Misc Major",
]

DAILY_EXPENSES = [
    "Groceries / Kirana",
    "Dining / Food / Swiggy",
    "Fuel / Petrol",
    "Recharge / Mobile",
    "Entertainment / OTT",
    "Transport / Uber / Auto",
    "Tea / Coffee / Snacks",
    "Household Items",
    "Personal Care",
    "Misc Daily",
]

INCOME_SOURCES = [
    "Salary / Main Income",
    "Freelance / Side Income",
    "Rental Income",
    "Investments / Dividends",
    "Other Income",
]

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# Map excel category names → website category names
CATEGORY_MAP = {
    "Rent / Home Loan": "Rent",
    "Electricity Bill": "Electricity Bill",
    "LIC Premium": "LIC Premium",
    "EMI / Loan Payment": "EMI / Loan",
    "School / Education": "Education",
    "Insurance (Other)": "Other",
    "Property Tax / Maintenance": "Other",
    "Internet / DTH Recharge": "Recharge / Mobile",
    "Groceries (Bulk)": "Groceries",
    "Groceries / Kirana": "Groceries",
    "Medical / Health": "Medical / Health",
    "Travel / Vacation": "Travel",
    "Shopping (Big)": "Shopping",
    "Family Support": "Friends / Transfer",
    "Misc Major": "Other",
    "Dining / Food / Swiggy": "Dining / Food",
    "Fuel / Petrol": "Fuel",
    "Recharge / Mobile": "Recharge / Mobile",
    "Entertainment / OTT": "Entertainment",
    "Transport / Uber / Auto": "Travel",
    "Tea / Coffee / Snacks": "Dining / Food",
    "Household Items": "Shopping",
    "Personal Care": "Shopping",
    "Misc Daily": "Other",
}

# Reverse map: website category → list of Excel category names (in priority order)
# First match = preferred Excel row for writing
WEBCAT_TO_EXCEL = {}
for excel_cat, web_cat in CATEGORY_MAP.items():
    if web_cat not in WEBCAT_TO_EXCEL:
        WEBCAT_TO_EXCEL[web_cat] = []
    WEBCAT_TO_EXCEL[web_cat].append(excel_cat)

# Reorder "Other" category: Misc first, then specific named rows
if "Other" in WEBCAT_TO_EXCEL:
    others = WEBCAT_TO_EXCEL["Other"]
    priority = [c for c in ["Misc Major", "Misc Daily"] if c in others]
    rest = [c for c in others if c not in priority]
    WEBCAT_TO_EXCEL["Other"] = priority + rest

# ─── Styles ──────────────────────────────────────────────────────────────────

if _HAS_OPENPYXL:
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ALTERNATE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    BORDER_THIN = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
else:
    HEADER_FONT = HEADER_FILL = ALTERNATE_FILL = BORDER_THIN = None
CATEGORY_COLORS = {
    "Electricity Bill": "FFC000",
    "LIC Premium": "FF6600",
    "Groceries": "70AD47",
    "Dining / Food": "FF0000",
    "Fuel": "4472C4",
    "Shopping": "7030A0",
    "EMI / Loan": "C00000",
    "Rent": "00B050",
    "Recharge / Mobile": "00B0F0",
    "Entertainment": "FF3399",
    "Travel": "ED7D31",
    "Medical / Health": "FF6699",
    "Education": "9966FF",
    "Friends / Transfer": "A5A5A5",
    "Other": "808080",
}

ALL_EXCEL_CATEGORIES = MAJOR_EXPENSES + DAILY_EXPENSES + INCOME_SOURCES


# ─── Balance Helpers (Start/End per month) ────────────────────

BALANCE_START_ROW = 2    # Row 2, column 6 (F) = start balance
BALANCE_END_ROW = 38     # Row 38 = end balance (after Grand Total at row 37)


def _write_balances_to_excel(wb, balance_data):
    """Write monthly start/end balances to Excel sheets.
    balance_data: dict of {month_name: {"start": float, "end": float}}
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    bold_font = Font(bold=True, size=10)

    for month_name in MONTH_NAMES:
        if month_name not in wb.sheetnames:
            continue
        ws = wb[month_name]
        bd = balance_data.get(month_name, {})

        # Start balance at F2
        ws.cell(row=2, column=6, value=bd.get("start", 0))
        ws.cell(row=2, column=6).number_format = '#,##0'
        ws.cell(row=2, column=5, value="Start Balance:")
        ws.cell(row=2, column=5).font = Font(bold=True, size=9)

        # End balance at row 38
        ws.cell(row=BALANCE_END_ROW, column=5, value="End Balance:")
        ws.cell(row=BALANCE_END_ROW, column=5).font = bold_font
        ws.cell(row=BALANCE_END_ROW, column=5).fill = green_fill
        ws.cell(row=BALANCE_END_ROW, column=5).alignment = Alignment(horizontal="right")
        end_val = bd.get("end", 0)
        ws.cell(row=BALANCE_END_ROW, column=6, value=end_val)
        ws.cell(row=BALANCE_END_ROW, column=6).number_format = '#,##0'
        ws.cell(row=BALANCE_END_ROW, column=6).font = bold_font
        ws.cell(row=BALANCE_END_ROW, column=6).fill = green_fill


def _read_balances_from_excel(wb):
    """Read monthly start/end balances from Excel sheets.
    Returns dict of {month_key (YYYY-MM): {"start": float, "end": float}}
    """
    month_names = MONTH_NAMES
    year = 2026
    if "start" in wb.sheetnames:
        yr = wb["start"]["E3"].value
        if yr:
            year = int(yr)

    result = {}
    for i, month_name in enumerate(month_names):
        if month_name not in wb.sheetnames:
            continue
        ws = wb[month_name]
        month_key = f"{year}-{i+1:02d}"
        start_val = ws.cell(row=2, column=6).value
        end_val = ws.cell(row=BALANCE_END_ROW, column=6).value
        try:
            start_val = float(start_val) if start_val else 0
        except (ValueError, TypeError):
            start_val = 0
        try:
            end_val = float(end_val) if end_val else 0
        except (ValueError, TypeError):
            end_val = 0
        result[month_key] = {"start": round(start_val, 2), "end": round(end_val, 2)}
    return result


def _find_excel_row(ws, excel_cat_name):
    """Find the row number in a month sheet for a given Excel category name."""
    for row_idx in range(6, 40):
        b_val = ws.cell(row=row_idx, column=2).value
        if b_val and str(b_val).strip() == excel_cat_name:
            return row_idx
    return None


def _get_excel_category_for(web_category, description=""):
    """Get the best Excel category name for a website category.
    Uses description as tiebreaker when multiple Excel categories map to same web category."""
    candidates = WEBCAT_TO_EXCEL.get(web_category, [])
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    # Multiple candidates — try to match by description
    desc_lower = description.lower()
    
    # First: exact category name in description (e.g. "Groceries / Kirana" in desc)
    for c in candidates:
        if c.lower() in desc_lower:
            return c
    
    # Second: distinctive single-word match (long words unique to one candidate)
    for c in candidates:
        words = [w for w in c.lower().replace("/", " ").split() if len(w) > 3]
        other_words = set()
        for oc in candidates:
            if oc != c:
                other_words.update(oc.lower().replace("/", " ").split())
        unique_words = [w for w in words if w not in other_words]
        for uw in unique_words:
            if uw in desc_lower:
                return c
    
    # Last resort: return first candidate
    return candidates[0]


def ensure_monthly_sheets(wb):
    """Ensure all 12 monthly sheets exist (create if missing)."""
    for m in MONTH_NAMES:
        if m not in wb.sheetnames:
            ws = wb.create_sheet(m)
            for day in range(1, 32):
                col = day + 3
                ws.cell(row=5, column=col, value=day).font = Font(bold=True, size=10)
            ws.cell(row=5, column=3, value="Budgeted").font = Font(bold=True, size=10)
            ws.column_dimensions["B"].width = 32


def read_transactions_from_excel():
    if not _HAS_OPENPYXL:
        return []
    """
    Read transactions from all 12 monthly calendar sheets.
    Returns flat list of transaction dicts.
    All cards default to "other" — user assigns cards on the website.
    """
    if not EXCEL_PATH.exists():
        return []

    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    except Exception:
        return []

    transactions = []
    year = 2026
    if "start" in wb.sheetnames:
        yr = wb["start"]["E3"].value
        if yr:
            year = int(yr)

    for month_idx, month_name in enumerate(MONTH_NAMES):
        if month_name not in wb.sheetnames:
            continue
        ws = wb[month_name]
        month_num = month_idx + 1

        for row_idx in range(6, ws.max_row + 1):
            b_val = ws.cell(row=row_idx, column=2).value
            if not b_val:
                continue
            cat_name = str(b_val).strip()
            if not cat_name or cat_name in ["None", "Day-To-Day Expenses", "Income", "Grand Total Expenses"]:
                continue
            if cat_name.startswith("="):
                continue

            if cat_name not in ALL_EXCEL_CATEGORIES:
                continue

            web_cat = CATEGORY_MAP.get(cat_name, "Other")
            card_id = "other"  # No auto-card — user picks on website
            txn_type = "debit"

            is_income = False
            for check_row in range(row_idx - 1, 5, -1):
                check_a = ws.cell(row=check_row, column=1).value
                if check_a and "Income" in str(check_a):
                    is_income = True
                    break
                if check_a and ("Major" in str(check_a) or "Day" in str(check_a)):
                    break

            if is_income:
                txn_type = "credit"
                web_cat = "Other"

            for day_col in range(4, 35):
                cell_val = ws.cell(row=row_idx, column=day_col).value
                if cell_val is None:
                    continue
                try:
                    amount = float(cell_val)
                except (ValueError, TypeError):
                    continue
                if amount == 0:
                    continue

                day_num = day_col - 3
                date_str = f"{year}-{month_num:02d}-{day_num:02d}"

                transactions.append({
                    "date": date_str,
                    "description": f"{cat_name}",
                    "amount": round(abs(amount), 2),
                    "category": web_cat if not is_income else "Other",
                    "card_id": card_id,
                    "txn_type": txn_type,
                    "notes": "",
                    "source": "excel",
                    "month": month_name,
                    "day": day_num,
                    "excel_cat": cat_name,
                })

    wb.close()
    return transactions


def write_transactions_to_excel(transactions):
    if not _HAS_OPENPYXL:
        return
    """Create/update a flat 'Transactions' sheet for website reference."""
    if not EXCEL_PATH.exists():
        return
    wb = openpyxl.load_workbook(str(EXCEL_PATH))

    if "Transactions" in wb.sheetnames:
        del wb["Transactions"]

    ws = wb.create_sheet("Transactions")
    headers = ["Date", "Description", "Amount", "Category", "Card", "Type", "Notes", "Cashback"]
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER_THIN
    widths = {"A": 14, "B": 40, "C": 12, "D": 18, "E": 16, "F": 10, "G": 30, "H": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    for i, txn in enumerate(transactions):
        row = i + 2
        ws.cell(row=row, column=1, value=txn["date"])
        ws.cell(row=row, column=2, value=txn["description"])
        ws.cell(row=row, column=3, value=txn["amount"])
        ws.cell(row=row, column=4, value=txn["category"])
        ws.cell(row=row, column=5, value=txn.get("card_id", "other"))
        ws.cell(row=row, column=6, value=txn.get("txn_type", "debit"))
        ws.cell(row=row, column=7, value=txn.get("notes", ""))
        ws.cell(row=row, column=8, value=txn.get("cashback", 0))
        if i % 2 == 1:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = ALTERNATE_FILL

    wb.save(str(EXCEL_PATH))
    wb.close()


def add_transaction_to_excel(txn):
    if not _HAS_OPENPYXL:
        return
    """
    Add a single transaction to the appropriate month sheet cell.
    Uses exact category matching via WEBCAT_TO_EXCEL reverse map.
    """
    if not EXCEL_PATH.exists():
        return

    wb = openpyxl.load_workbook(str(EXCEL_PATH))

    try:
        dt = datetime.strptime(txn["date"], "%Y-%m-%d")
    except (ValueError, TypeError):
        dt = datetime.now()
    month_name = MONTH_NAMES[dt.month - 1]
    day_num = dt.day

    if month_name in wb.sheetnames:
        ws = wb[month_name]

        # Determine Excel category name from website category + description
        web_cat = txn.get("category", "")
        excel_cat_name = _get_excel_category_for(web_cat, txn.get("description", ""))

        if not excel_cat_name:
            # Fallback: try to find by description
            for candidate in ALL_EXCEL_CATEGORIES:
                if candidate.lower() in txn.get("description", "").lower():
                    excel_cat_name = candidate
                    break

        if excel_cat_name:
            found_row = _find_excel_row(ws, excel_cat_name)
            if found_row:
                col = day_num + 3
                existing = ws.cell(row=found_row, column=col).value
                new_val = (float(existing) if existing else 0) + txn["amount"]
                ws.cell(row=found_row, column=col, value=round(new_val, 2))
    else:
        ensure_monthly_sheets(wb)

    wb.save(str(EXCEL_PATH))
    wb.close()

    # Also update flat Transactions sheet — write from DB directly to preserve card IDs
    wb = None
    try:
        import sqlite3
        conn = sqlite3.connect(str(BASE_DIR / "finance.db"))
        db_rows = conn.execute(
            "SELECT id, date, description, amount, category, card_id, txn_type, notes, cashback "
            "FROM transactions ORDER BY date, id"
        ).fetchall()
        conn.close()
        all_txns = []
        for r in db_rows:
            txn_id, date_str, desc, amount, category, card_id, txn_type, notes, cashback = r
            try:
                dt = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
            except (ValueError, TypeError):
                continue
            month_name = MONTH_NAMES[dt.month - 1]
            all_txns.append({
                "date": str(date_str)[:10],
                "description": desc or "",
                "amount": round(abs(amount), 2),
                "category": category or "",
                "card_id": card_id or "other",
                "txn_type": txn_type or "debit",
                "notes": notes or "",
                "cashback": round(cashback or 0, 2),
                "month": month_name,
            })
        write_transactions_to_excel(all_txns)
    except Exception as e:
        print(f"⚠️ Could not update Transactions sheet: {e}")


def excel_sync_to_db():
    if not _HAS_OPENPYXL:
        return 0
    """Sync Excel monthly sheets → SQLite database, preserving card assignments."""
    transactions = read_transactions_from_excel()

    import sqlite3
    conn = sqlite3.connect(str(BASE_DIR / "finance.db"))
    conn.row_factory = sqlite3.Row

    # Save existing card assignments AND person data for Excel-sourced transactions
    existing_data = {}
    for row in conn.execute("SELECT date, description, amount, category, card_id, person FROM transactions WHERE source='excel'"):
        key = (row["date"], row["description"], row["amount"], row["category"])
        existing_data[key] = {"card_id": row["card_id"], "person": row["person"]}

    conn.execute("DELETE FROM transactions WHERE source='excel'")

    inserted = 0
    for txn in transactions:
        key = (txn["date"], txn["description"], txn["amount"], txn["category"])
        saved = existing_data.get(key, {})
        card_id = saved.get("card_id", txn.get("card_id", "other"))
        person = saved.get("person", txn.get("person", ""))
        conn.execute(
            "INSERT INTO transactions (date, description, amount, category, card_id, txn_type, notes, source, person) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, 'excel', ?)",
            (txn["date"], txn["description"], txn["amount"],
             txn["category"], card_id, txn["txn_type"], txn.get("notes", ""), person)
        )
        inserted += 1
    conn.commit()
    conn.close()
    return inserted


def db_sync_to_excel():
    if not _HAS_OPENPYXL:
        return 0
    """
    Sync DB → Excel (write to monthly sheets).
    First clears all data cells, then writes fresh from DB.
    Category matching uses exact Excel category names.
    """
    import sqlite3
    conn = sqlite3.connect(str(BASE_DIR / "finance.db"))
    rows = conn.execute(
        "SELECT id, date, description, amount, category, card_id, txn_type, notes, cashback "
        "FROM transactions ORDER BY date, id"
    ).fetchall()
    conn.close()

    if not EXCEL_PATH.exists():
        print("⚠️ Excel file not found, cannot sync")
        return 0

    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH))
    except PermissionError:
        print("⚠️ Excel file is open/locked — sync skipped. Close Excel and try again.")
        return 0
    except FileNotFoundError:
        print("⚠️ Excel file not found")
        return 0

    # Clear all numeric data cells in month sheets first
    for month_name in MONTH_NAMES:
        if month_name not in wb.sheetnames:
            continue
        ws = wb[month_name]
        for row_idx in range(6, 40):
            for col in range(4, 35):
                cell = ws.cell(row=row_idx, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.value = None

    # Group DB transactions by (month, excel_category, day) and sum amounts
    from collections import defaultdict
    grouped = defaultdict(float)
    txn_info = {}  # (month, excel_cat, day) -> first txn details

    for r in rows:
        txn_id, date_str, desc, amount, category, card_id, txn_type, notes, cashback = r
        try:
            dt = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
        except (ValueError, TypeError):
            continue
        month_name = MONTH_NAMES[dt.month - 1]
        day_num = dt.day

        # Map website category → Excel category name
        excel_cat_name = _get_excel_category_for(category or "", desc or "")
        if not excel_cat_name:
            # Fallback: try description match with keywords
            for candidate in ALL_EXCEL_CATEGORIES:
                if candidate.lower() in (desc or "").lower():
                    excel_cat_name = candidate
                    break
        if not excel_cat_name:
            continue

        key = (month_name, excel_cat_name, day_num)
        grouped[key] += abs(amount)
        if key not in txn_info:
            txn_info[key] = {
                "desc": desc, "category": category, "card_id": card_id,
                "txn_type": txn_type, "notes": notes,
            }

    written = 0
    for (month_name, excel_cat_name, day_num), total_amount in grouped.items():
        if month_name not in wb.sheetnames:
            continue
        ws = wb[month_name]
        found_row = _find_excel_row(ws, excel_cat_name)
        if found_row:
            col = day_num + 3
            ws.cell(row=found_row, column=col, value=round(total_amount, 2))
            written += 1

    # Write balances to Excel
    try:
        bal_conn = sqlite3.connect(str(BASE_DIR / "finance.db"))
        bal_conn.row_factory = sqlite3.Row
        bal_rows = bal_conn.execute("SELECT month, start_balance, end_balance FROM monthly_balances").fetchall()
        bal_conn.close()
        balance_data = {}
        for br in bal_rows:
            m = br["month"]
            parts = m.split("-")
            if len(parts) == 2:
                mon = int(parts[1])
                if 1 <= mon <= 12:
                    month_name = MONTH_NAMES[mon - 1]
                    balance_data[month_name] = {
                        "start": br["start_balance"] or 0,
                        "end": br["end_balance"] or 0,
                    }
        _write_balances_to_excel(wb, balance_data)
    except Exception as e:
        print(f"⚠️ Could not write balances: {e}")

    try:
        wb.save(str(EXCEL_PATH))
    except PermissionError:
        print("⚠️ Could not save Excel — file is open. Close Excel first.")
        wb.close()
        return written
    wb.close()

    # Also update flat Transactions sheet — write from DB directly to preserve card IDs
    all_txns = []
    for r in rows:
        txn_id, date_str, desc, amount, category, card_id, txn_type, notes, cashback = r
        try:
            dt = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
        except (ValueError, TypeError):
            continue
        month_name = MONTH_NAMES[dt.month - 1]
        excel_cat_name = _get_excel_category_for(category or "", desc or "")
        if not excel_cat_name:
            for candidate in ALL_EXCEL_CATEGORIES:
                if candidate.lower() in (desc or "").lower():
                    excel_cat_name = candidate
                    break
        if not excel_cat_name:
            excel_cat_name = "Misc Major"
        all_txns.append({
            "date": str(date_str)[:10],
            "description": desc or "",
            "amount": round(abs(amount), 2),
            "category": category or "",
            "card_id": card_id or "other",
            "txn_type": txn_type or "debit",
            "notes": notes or "",
            "cashback": round(cashback or 0, 2),
            "month": month_name,
            "excel_cat": excel_cat_name,
        })
    write_transactions_to_excel(all_txns)

    return written


def smart_sync():
    if not _HAS_OPENPYXL:
        return
    """
    Two-way smart sync.
    Count-based: the side with more transactions wins.
    Equal-count: compare totals — if DB total differs from Excel total,
    someone modified amounts. Prefer the side with a larger total
    (indicates additions/modifications). If totals match, prefer DB→Excel
    (website-first approach since web is the primary edit interface).
    """
    excel_txns = read_transactions_from_excel()
    db_txns = _get_db_transactions()

    excel_count = len(excel_txns)
    db_count = len(db_txns)

    if excel_count > db_count:
        # Excel has more data → user added in Excel
        count = excel_sync_to_db()
        return "excel_to_db", count
    elif db_count > excel_count:
        # DB has more data → user added on website
        count = db_sync_to_excel()
        return "db_to_excel", count
    else:
        # Same count — compare total amounts to detect modifications
        excel_total = sum(abs(t.get("amount", 0)) for t in excel_txns)
        db_total = sum(abs(t["amount"]) for t in db_txns) if db_txns else 0
        if db_total != excel_total:
            # Totals differ — someone modified amounts
            if db_total > excel_total:
                # DB has more value → website changes
                count = db_sync_to_excel()
                return "db_to_excel", count
            else:
                # Excel has more value → Excel changes
                count = excel_sync_to_db()
                return "excel_to_db", count
        # Totals match too — prefer DB→Excel (website-first)
        count = db_sync_to_excel()
        return "db_to_excel", count


def _get_db_transactions():
    import sqlite3
    conn = sqlite3.connect(str(BASE_DIR / "finance.db"))
    conn.row_factory = sqlite3.Row
    # Ensure tables exist (could be first run on a fresh DB)
    conn.execute("CREATE TABLE IF NOT EXISTS transactions (id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT NOT NULL, description TEXT NOT NULL, amount REAL NOT NULL, category TEXT DEFAULT 'Other', card_id TEXT DEFAULT 'other', txn_type TEXT DEFAULT 'debit', notes TEXT DEFAULT '', source TEXT DEFAULT 'manual', created_at TEXT DEFAULT (datetime('now','localtime')), person TEXT DEFAULT \"\", cashback REAL DEFAULT 0, user_id INTEGER)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_txn_date ON transactions(date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_txn_user ON transactions(user_id)")
    rows = conn.execute(
        "SELECT date, description, amount, category, card_id, txn_type, notes "
        "FROM transactions ORDER BY date DESC, id DESC"
    ).fetchall()
    conn.close()
    return [{k: r[k] for k in r.keys()} for r in rows]


def rename_excel_for_year(year):
    new_path = BASE_DIR / f"expense_tracker_{year}.xlsx"
    copy2(str(EXCEL_PATH), str(new_path))
    return new_path


# ─── Initialize ──────────────────────────────────────────────────────────────

def init_excel():
    """Create fresh Excel if missing using the monthly sheets creator."""
    if EXCEL_PATH.exists():
        return True

    print("📦 Creating fresh Excel with monthly sheets...")
    import subprocess
    result = subprocess.run(
        ["python3", str(BASE_DIR / "create_monthly_excel.py"), str(EXCEL_PATH)],
        capture_output=True, text=True
    )
    print(result.stdout.strip())
    if result.returncode != 0:
        print(f"❌ Error: {result.stderr.strip()}")
        return False
    return True


if __name__ == "__main__":
    init_excel()
    txns = read_transactions_from_excel()
    print(f"📊 Monthly sheets parsed: {len(txns)} transactions found")
    for t in txns[:10]:
        print(f"  {t['date']} | {t['description'][:35]:35s} | ₹{t['amount']:>8.2f} | {t['category']}")
    if len(txns) > 10:
        print(f"  ... and {len(txns)-10} more")
